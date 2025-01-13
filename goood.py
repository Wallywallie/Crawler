import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import time
from selenium.webdriver.common.by import By
from typing import(List, Tuple)
import logging
from abc import ABC, abstractmethod
import re, os
import pickle
import openpyxl
import json

# 配置日志
logging.basicConfig(level=logging.INFO, 
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    handlers=[
                        logging.FileHandler("search_results.log"),
                        logging.StreamHandler()  # 同时输出到控制台
                    ])
logger = logging.getLogger(__name__)


class Content:
    """
    表示内容，可以是段落或图片。

    Attributes:
        content_type (str): 内容类型 ('paragraph' 或 'image')。
        content (str): 内容（段落文字或图片链接）。
        image_caption (str): 图片说明（如果内容类型是 'image'）。
    """    
    def __init__(self, type: str, content: str, image_caption: str = None):
        self.type = type  # 'paragraph' or 'image'
        self.content = content
        self.image_caption = image_caption

    def is_para(self):
        return self.type == 'paragraph'
    
    def get(self):
        return self.content
    
    def empty(self) -> bool:
        if self.content:
            return False
        return True


class Case:
    """
    表示一个案例，包含标题和多个内容项。

    Attributes:
        title (str): 案例的标题。
        contents (list): 案例中的内容项列表。
    """    
    def __init__(self, title: str):
        self.title = title
        self.contents = []
        self.source = ''
        self.key_word = ''

    def add_content(self, content: Content):
        self.contents.append(content)

    def print_case(self):
        for i in self.contents:
            print(i.content)
            print("")

    def save_all(self, file_name:str, folder_name:str = "data"):

        sanitized_name =  self.__sanitize_name(file_name)
        sanitized_name = sanitized_name + '.txt'
        subfolder_path = self.__subfolder(folder_name)
        path = os.path.join(subfolder_path, sanitized_name)

        paragraphs = '\n'.join(i.content for i in self.contents)
        print(paragraphs)
        with open(path, 'w', encoding='utf-8') as f:
            f.write(paragraphs)

    def __sanitize_name(self, file_name: str) -> str :

        # 移除或替换不允许的特殊字符，保留字母、数字、空格、连字符、下划线
        sanitized_name = re.sub(r'[\/:*?"<>|]', '_', file_name)
    
        # 如果文件名过长，进行截断
        max_length = 255 - len('.txt')  # 保证文件名+扩展名不超过255个字符
        if len(sanitized_name) > max_length:
            sanitized_name = sanitized_name[:max_length] 

        return sanitized_name

    def __subfolder(self, folder_name : str)->str:
        current_path = os.getcwd()
        subfolder_path = os.path.join(current_path, folder_name)
        if not os.path.exists(subfolder_path):
            os.makedirs(subfolder_path)
        return subfolder_path


    def save_text(self, file_name:str, folder_name:str = "data"):
        sanitized_name =  self.__sanitize_name(file_name)
        sanitized_name = sanitized_name + '.txt'
        subfolder_path = self.__subfolder(folder_name)
        path = os.path.join(subfolder_path, sanitized_name)
        paragraphs = '\\n'.join(i.content for i in self.contents if i.type == "paragraph")
        paragraphs = self.title + '\\n' + paragraphs
        with open(path, 'w', encoding='utf-8') as f:
            f.write(paragraphs)  


    def serialize(self, file_name:str, folder_name:str = 'obj') :
        sanitized_name =  self.__sanitize_name(file_name)
        sanitized_name = sanitized_name + '.pkl'        
        subfolder_path = self.__subfolder(folder_name)
        path = os.path.join(subfolder_path, sanitized_name)

        with open(path, 'wb') as f :     
            pickle.dump(self, f)   

    @staticmethod
    def deserialize(folder_name:str = 'obj') :
        current_path = os.getcwd()
        subfolder_path = os.path.join(current_path, folder_name)

        # 遍历子文件夹中的所有文件
        for root, dirs, files in os.walk(subfolder_path):
            for file_name in files:
                file_path = os.path.join(root, file_name)
                if file_name.endswith('.pkl'): 
                    logger.info(f"Processing file: {file_path}")
                    try:
                        with open(file_path, 'rb') as f:
                            obj = pickle.load(f)  # 反序列化对象
                            obj.save_text(obj.title)
       
                    except Exception as e:
                        logger.error(f"Error deserializing {file_path}: {e}")

# 定义抽象策略类
class CrawlStrategy(ABC): 
    @abstractmethod
    def get_search_results(self, query):
        pass

    @abstractmethod
    def extract_info(url):
        pass


class GdCrawlStrategy(CrawlStrategy):

    def get_search_results(driver : webdriver.Chrome,query:str, page:int = 1, wait_time = 1)-> Tuple[List[str], List[str]]:
        """
        获取搜索结果链接列表。

        Args:
        query (str): 搜索关键词。
        page (int): 页码，默认为1。
        wait_time : 等待页面加载的时间，默认为1秒。

        Returns:
        List[str]: 包含搜索结果链接的列表。
        """
        results = []
        titles = []
        base_url = "https://www.gooood.cn/search/"
        url = f"{base_url}{query}/page/{page}"

        try:

            driver.get(url)
            time.sleep(wait_time)  # 等待页面加载
            soup = BeautifulSoup(driver.page_source, 'html.parser')
    
            links = soup.find_all("div", class_ = 'post-thumbnail')

            
            for i in links:
                a_tag = i.find("a")
                if a_tag is None:
                    continue
                results.append(a_tag['href'])
                title = a_tag.get_text(strip=True)
                titles.append(title)
                logger.info(f"Found link: {a_tag['href']}")
        except Exception as e:
            logger.error(f"An error occurred in get_search_results: {e}")   
               

        logger.info(f"Total links found: {len(results)}")
        return results, titles
    
    def extract_info(driver : webdriver.Chrome, url:str, wait_time = 1 ) -> Case:
        """
        从给定的URL提取案例信息。

        Args:
        url (str): 要提取信息的网页URL。
        wait_time : 等待页面加载的时间，默认为1秒。

        Returns:
        Case: 包含提取内容的Case对象。
        """

        try:
        
            driver.get(url)
            time.sleep(wait_time)  # 等待页面加载
            
            soup = BeautifulSoup(driver.page_source, 'html.parser')

            # get title | cover | content | picture | caption
            title = soup.find('title').get_text() if soup.find('title') else "No Title"
            #cover = soup.find('meta', property="og:image")['content']
            case = Case(title)

            post_content = soup.find('post-content')
            if post_content:
                p_tags = post_content.find_all('p')
                if p_tags:
                    for p in p_tags:
                        if p.get_text(strip=True):
                            case.add_content(Content('paragraph',p.get_text(strip=True) ))
                        a_tag = p.find('a')
                        if a_tag:
                            img_tag = a_tag.find('img')
                            if img_tag:
                                case.add_content(Content('image',img_tag['src']))
        except Exception as e:
            logger.error(f"An error occurred in extract_info: {e}")

        logging.info(f"Extraction completed for URL: {url}")

        return case


class ArchDailyStrategy(CrawlStrategy):

    def get_search_results(driver : webdriver.Chrome,query:str, page:int = 1, wait_time = 1)->Tuple[List[str], List[str]]:
        """
        获取搜索结果链接列表。

        Args:
        query (str): 搜索关键词。
        page (int): 页码，默认为1。
        wait_time : 等待页面加载的时间，默认为1秒。

        Returns:
        List[str]: 包含搜索结果链接的列表。
        """
        results = []
        titles = []
        base_url = "https://www.archdaily.cn/search/api/v1/cn/all?q="
        
        url = f"{base_url}{query}&page={page}"
        try:

            driver.get(url)
            time.sleep(wait_time)  # 等待页面加载
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            pre_tag = soup.find("pre")
            data = json.loads(pre_tag.text)
            for i in data["results"]:
                url = i["url"]
                titles.append(i["title"])
                results.append(url)
                logger.info(f"Found link: {url}")

        except Exception as e:
            logger.error(f"An error occurred in get_search_results: {e}")

        return results,titles  
    

    def extract_info(driver : webdriver.Chrome, url:str, wait_time = 1 )-> Case:
        """
        从给定的URL提取案例信息。

        Args:
        url (str): 要提取信息的网页URL。
        wait_time : 等待页面加载的时间，默认为1秒。

        Returns:
        Case: 包含提取内容的Case对象。
        """
        print(url)
        try:
        
            driver.get(url)
            time.sleep(wait_time)  # 等待页面加载
            
            soup = BeautifulSoup(driver.page_source, 'html.parser')


            title = soup.find('title').get_text() if soup.find('title') else "No Title"
            case = Case(title)

            
            start_content = soup.find("span", class_ = "less-txt hide")
            pre = start_content.find_next().find_previous_sibling()
            for sib in pre.next_siblings:
                if sib.name == "figure":
                    img = sib.find("a")["data-image"]
                    content = Content("image", img)
                    case.add_content(content)
                elif sib.name == "p":
                    para = sib.get_text(strip=True)
                    if para:
                        content = Content("paragraph", para)
                        case.add_content(content)

        except Exception as e:
            logger.error(f"An error occurred in extract_info: {e}")

        logging.info(f"Extraction completed for URL: {url}")

        return case 


class YouFang(CrawlStrategy):
    def get_search_results(driver : webdriver.Chrome, query:str, page:int = 1, wait_time = 1)-> Tuple[List[str], List[str]]:
        results = []
        titles = []
        base_url = "https://www.archiposition.com/page/"
        url = f"{base_url}{page}?s={query}&cat=4" #cat=4 显示文章 cat=all显示所有

        try:

            driver.get(url)
            time.sleep(wait_time)  # 等待页面加载
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            div_tags = soup.find_all("h2", class_="article-title")
            for div in div_tags:
                a_tag = div.find("a")
                if a_tag and 'href' in a_tag.attrs:
                    link = "https://www.archiposition.com" + a_tag['href']
                    title = a_tag.get_text(strip=True)
                    results.append(link)
                    titles.append(title)
                    logger.info(f"Found link: {link}")
            
        except Exception as e:
            logger.error(f"An error occurred in get_search_results: {e}")   
               

        logger.info(f"Total links found: {len(results)}")
        return results,titles

        
    
    def extract_info(driver : webdriver.Chrome, url:str, wait_time = 1 )-> Case:
        try:
            driver.get(url)
            time.sleep(wait_time)  # 等待页面加载
            
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            title = soup.find('title').get_text() if soup.find('title') else "No Title"
            case = Case(title)
            div_tag = soup.find("div", class_="detail-left rich-text")
            div = div_tag.find("div", class_="detail-content")
            if div:
                for child in div.find_all(['p', 'figure']):
                    if child.name == 'p':
                        # 创建 Content 对象并设置 type 和 content
                        content_obj = Content(type='paragraph', content=child.get_text(strip=True))
                        case.add_content(content_obj)
                    
                    elif child.name == 'figure':
                        img_tag = child.find('img')
                        figcaption_tag = child.find('figcaption')
                        if img_tag and 'src' in img_tag.attrs:
                            content_text = img_tag['src']
                            caption_text = figcaption_tag.get_text(strip=True) if figcaption_tag else None
                            content_obj = Content(type='image', content=content_text, image_caption=caption_text)
                            case.add_content(content_obj)


        except Exception as e:
            logger.error(f"An error occurred in extract_info: {e}")

        logging.info(f"Extraction completed for URL: {url}")        
        return case

# 爬虫类
class Crawler:

    def create_driver(self) -> webdriver.Chrome:
        """
        创建并配置一个新的 Chrome WebDriver 实例。
        
        返回:
        webdriver.Chrome: 配置好的 WebDriver 实例。
        """
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        driver = webdriver.Chrome(executable_path=ChromeDriverManager().install(), options=options)
        return driver    


    def __init__(self, strategy: CrawlStrategy):
        self.strategy = strategy
        self.driver = self.create_driver()

    def set_strategy(self, strategy: CrawlStrategy):
        self.strategy = strategy

    def get_search_results(self, query, page = 1,wait_time=1):
  
        return self.strategy.get_search_results(self.driver,query,page,wait_time)
    
    def close(self):
        self.driver.quit()
        logger.info("Browser closed.")

    def extract_info(self, url, wait_time=1):
        return self.strategy.extract_info(self.driver, url, wait_time)    





from difflib import SequenceMatcher

def is_fuzzy_match(search_term, title, threshold=0.5) -> bool:
    # 使用 SequenceMatcher 计算相似度
    matcher = SequenceMatcher(None, search_term, title)
    match_ratio = matcher.find_longest_match(0, len(search_term), 0, len(title)).size / len(search_term)
    print(title,">>", match_ratio)
    
    # 判断相似度是否超过阈值
    return match_ratio >= threshold


"""
query = "尼康直营店"
crawler = Crawler(ArchDailyStrategy)
urls = crawler.get_search_results(query)
for i in  urls:
    case = crawler.extract_info(i,wait_time=0.7)
    case.serialize(case.title)

crawler.set_strategy(GdCrawlStrategy)

urls = crawler.get_search_results(query)
for i in  urls:
    case = crawler.extract_info(i,wait_time=0.7)
    case.serialize(case.title)

crawler.close()
"""

def crawl_feat():
    query_lst_1 = ["工业", "厂房", "仓库", "车间", "工业遗产"]
    query_lst_2 = ["公共", "教育" , "办公", "商业","酒店","民宿", "金融", "剧院", "展览", "美术", "博物", "医疗", "疗养", "体育", "交通", "教堂", "纪念" ]
    query_lst_3 = ["居住", "住宅" ]
    query = query_lst_1 + query_lst_2 + query_lst_3
    strategy = [YouFang, GdCrawlStrategy, ArchDailyStrategy ]
    tag = ["yf", "gd", "arch"]

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    crawler = Crawler(YouFang)
    titles = []
    for j in range(0, len(strategy)):
        crawler.set_strategy(strategy[j])
        for q in query:
            urls,titles = crawler.get_search_results(q)
            for i in  urls:
                case = crawler.extract_info(i,wait_time=0.7)
                case.serialize(case.title,'feature')
                break
        break

    crawler.close()
        
def gen_title():
    import train
    import pandas as pd
    from tqdm import tqdm

    file_name = "overlap_arch.xlsx"
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    df = pd.read_excel(file_name, sheet_name="Sheet")  # 可以指定 sheet_name


    for index, row in tqdm(df.iterrows(), total=df.shape[0]):
        data = row.iloc[0]
        if type(data) is not str:
            continue
        res = train.extract_entity(data)
        sheet.cell(row=index + 2, column=4).value = res

    workbook.save(file_name)

def crawl():

    query_lst = ["酒店", "办公", "展览馆", "住宅 公寓", "工业 厂房"]
    strategy = [YouFang, GdCrawlStrategy, ArchDailyStrategy ]
    tag = ["yf", "gd", "arch"]


    crawler = Crawler(ArchDailyStrategy)
    titles = []

    file_name = "overlap_arch.xlsx"

    for q in range(len(query_lst)):
        query = query_lst[q]
        if os.path.exists(file_name):
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active    
            
        page = 1
        while sheet.max_row < (q + 1 ) * 200:
            urls,titles = crawler.get_search_results(query, page)
            print(f"searching: {query}, page: {page}, last row: {sheet.max_row}, limit: {(q + 1 ) * 200}")
            page = page + 1   
            last_row = sheet.max_row
            for index, value in  enumerate(titles, start=1):# start=1 表示从 A1 开始
                sheet.cell(row=last_row + index, column=1, value=value)
                sheet.cell(row=last_row + index, column=2, value=q)
                sheet.cell(row=last_row + index, column=3, value="arch")
            workbook.save(file_name) 
    crawler.close()

def process_excel(file_name: str):
    import pandas as pd
    import tqdm as tqdm

    df = pd.read_excel(file_name, sheet_name="Sheet1")

    stats = {
        "yf": {i: 0 for i in range(5)},  # 用字典初始化计数器
        "arch": {i: 0 for i in range(5)},
        "gd": {i: 0 for i in range(5)}   # 这里假设 gd 只有 0-3 的分类
    }

    # 总计变量
    total = {
        "yf": 0,
        "arch": 0,
        "gd": 0
    }


    # 遍历每一行的数据
    for index, row in df.iterrows():
        category = row.iloc[2]  # 第三列的值作为 category
        key = row.iloc[1]  # 第二列的值作为 key
        
        if category in stats:  # 判断是否为需要统计的类别
            total[category] += 1  # 更新总计
            if key in stats[category]:
                stats[category][key] += 1  # 更新分类计数器

    # 打印统计结果
    for category in total:
        print(f"Total {category}: {total[category]}")
        for key in stats[category]:
            print(f"{category}_{key}: {stats[category][key]}")


def get_training_obj():
    query_lst_1 = ["工业", "厂房", "仓库", "车间", "工业遗产"]
    query_lst_2 = ["公共", "教育" , "办公", "商业","酒店","民宿", "金融", "剧院", "展览", "美术", "博物", "医疗", "疗养", "体育", "交通", "教堂", "纪念" ]
    query_lst_3 = ["居住", "住宅" ]
    query = query_lst_1 + query_lst_2 + query_lst_3
    crawler = Crawler(YouFang)
    for q in query:
        urls, titles = crawler.get_search_results(q,2)
        for i in  urls:
            case = crawler.extract_info(i,wait_time=0.7)
            case.source = 'yf'
            case.key_word = q
            case.serialize(case.title)

    crawler.close()

def is_eng(para: str) -> bool:
    pattern = r'^[a-zA-Z0-9\s.,!?\'\"\-\(\): ;’—“”çé–]*$'  # 只允许英文字母、空格和常见标点符号
    return bool(re.match(pattern, para))

def write_excel(folder_name: str = 'obj'):

    workbook = openpyxl.Workbook()
    sheet = workbook.active 

    data = []
    current_path = os.getcwd()
    subfolder_path = os.path.join(current_path, folder_name)

    # 遍历子文件夹中的所有文件
    for root, dirs, files in os.walk(subfolder_path):
        for file_name in files:
            file_path = os.path.join(root, file_name)
            if file_name.endswith('.pkl'): 
                with open(file_path, 'rb') as f:
                    obj = pickle.load(f)  # 反序列化对象
                    data.append(obj)
    

    excel_name = "training_zh.xlsx"

    index = 1
    for i in data:
        paragraphs = '\\n'.join(ctt.content for ctt in i.contents if ctt.type == "paragraph" and not is_eng(ctt.content))
        paragraphs = i.title + '\\n' + paragraphs
        sheet.cell(row=index, column=1).value = i.title
        sheet.cell(row=index, column=2).value = paragraphs
        sheet.cell(row=index, column=3).value = i.key_word
        sheet.cell(row=index, column=3).value = i.source
        index += 1
   
 

    
    workbook.save(excel_name) 

def instru(file_name: str):
    import pandas as pd
    df = pd.read_excel(file_name, sheet_name="Sheet1")
    key = ["arch_name","arch_loc","architect","arch_com","built_time","arch_area", "site_area", "material"]
    data = []

    for index, row in df.iterrows():
        text = row.iloc[1]
        para = text.split('\\n')
        input = '\n'.join(para)
        print(input)
        output = {}
        for i in range(3, 11):

            output[key[i-3]] = row.iloc[i] if type(row.iloc[i]) == str else ""
        dic = {"instruction": """你是一个案例信息提取专家，你要从中提取案例描述的建筑作品的 建筑名称；建筑地址；建筑师；设计单位；建成时间；建筑面积；场地面积；材料. 以 json 格式输出.注意: 1. 输出的每一行都必须是正确的 json 字符串. 2. 没有找到对应信息时, 输出空字符串.""",
            "input":input,
            "output":output}
        data.append(dic)

    output_file = "data_1215.jsonl"
    with open(output_file, "w", encoding="utf-8") as file:
        for record in data:
            json_line = json.dumps(record, ensure_ascii=False)  # 确保非 ASCII 字符正常写入
            file.write(json_line + "\n")       

instru("training_cpl.xlsx")


def get_seceond_data(file_name: str):
    import pandas as pd
    df = pd.read_excel(file_name, sheet_name="Sheet")   
    data = []

    for index, row in df.iterrows():
        text = row.iloc[1]
        para = text.split('\\n')
        input = '\n'.join(para)
          
        dic = {"instruction": """你是一个案例信息提取专家，你要从中提取案例描述的建筑作品的 建筑名称；建筑地址；建筑师；设计单位；建成时间；建筑面积；场地面积；材料. 以 json 格式输出.注意: 1. 输出的每一行都必须是正确的 json 字符串. 2. 没有找到对应信息时, 输出空字符串.""",
            "input":input,
            "output":""}
        data.append(dic)
    output_file = "data_scd.jsonl"
    with open(output_file, "w", encoding="utf-8") as file:
        for record in data:
            json_line = json.dumps(record, ensure_ascii=False)  # 确保非 ASCII 字符正常写入
            file.write(json_line + "\n")     

def parse_string(input_str):
    import ast

    try:
        return ast.literal_eval(input_str)
    except (ValueError, SyntaxError):

        return None


def result_to_excel(file_name : str):
    import ast
    dict_list = []
    with open('data_scd.jsonl', 'r', encoding='utf-8') as f:
        for line in f:
            # 每一行是一个 JSON 对象
            json_obj = json.loads(line.strip()) 
            dict_list.append(parse_string(json_obj))
    
    
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active    

    print(len(dict_list))
    for i in range(34, len(dict_list)):
        data = dict_list[i]
        
        if type(data) == dict:
            try:
                sheet.cell(row=i + 2, column=4).value = data['arch_name']
                print(sheet.cell(row=i + 2, column=4).value)
                sheet.cell(row=i + 2, column=5).value = data['arch_loc']
                sheet.cell(row=i + 2, column=6).value = data['architect']
                sheet.cell(row=i + 2, column=7).value = data['arch_com']
                sheet.cell(row=i + 2, column=8).value = data['built_time']
                sheet.cell(row=i + 2, column=9).value = data['arch_area']
                sheet.cell(row=i + 2, column=10).value = data['site_area']
                sheet.cell(row=i + 2, column=11).value = data['material']
            except(KeyError):
                continue

    workbook.save('temp.xlsx')
            

#result_to_excel('training.xlsx')
